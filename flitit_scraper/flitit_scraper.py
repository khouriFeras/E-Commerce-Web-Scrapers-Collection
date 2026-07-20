#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
flitit_scraper.py

Scrapes product Description + Image URLs from flitit.com for rows in an
Excel catalog file, matching each row via Item Name only (barcode
matching disabled).

IMPORTANT: this script edits the workbook with openpyxl DIRECTLY on the
original file's cells (never through pandas read/write). This is
required because the source file has embedded product photos in the
"Picture" column -- a pandas read_excel()/to_excel() round-trip would
silently drop every embedded image and all formatting. openpyxl only
touches the specific cells we write to, so everything else (images,
colors, column widths, merged cells...) stays exactly as it was.

Before processing starts, any trailing "phantom" rows -- rows that have
leftover cell formatting (borders/colors/etc.) but no actual Item Name
-- are physically deleted. Excel/openpyxl can report ws.max_row far
beyond the real last data row just because formatting was applied down
there at some point; without this cleanup step the script would loop
(silently) over hundreds of empty rows.

OUTPUT: two files, produced by copying the processed workbook and then
HIDING the rows that don't belong in each copy (instead of deleting
them). Hiding -- not deleting -- is what keeps embedded per-row images
correctly aligned with their row.

    <basename>_found.xlsx     -> only MATCHED rows visible
    <basename>_notfound.xlsx  -> everything else visible (NO_RESULT / ERROR)

Usage:
    python flitit_scraper.py "C:\\path\\to\\file.xlsx"
    python flitit_scraper.py "C:\\path\\to\\file.xlsx" all
    (or just: python flitit_scraper.py   -- uses INPUT_FILE below)
"""

import os
import re
import shutil
import sys
import urllib.parse

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# =========================================================
# CONFIG
# =========================================================
INPUT_FILE = r"C:\JafarShop\E-Commerce-Web-Scrapers-Collection\flitit_scraper\Ambasador 2026.7.3.xlsx"

# Set to None to process ALL data rows. Currently limited to 10 rows for
# testing -- just change this value (or set to None) to remove/adjust it.
ROW_LIMIT = 10

BASE_URL = "https://flitit.com"
SEARCH_URL_TEMPLATE = "https://flitit.com/search?q={query}&options%5Bprefix%5D=last"

NAV_TIMEOUT = 30000  # ms

# Column headers we're looking for in the sheet.
# Barcode matching is disabled -- BARCODE_HEADER stays None everywhere,
# and only ITEM_NAME_HEADER is used to search/match products.
BARCODE_HEADER = None
ITEM_NAME_HEADER = "SKU"

# New columns we append.
NEW_COLUMNS = ["Description", "Image URLs", "Status"]

DESCRIPTION_SELECTOR = "div#tab1"

# Only main-content product links -- excludes header/footer nav links.
FIRST_PRODUCT_LINK_SELECTOR = 'main a[href*="/products/"]'

# Selectors scoped to the actual product media gallery only (never the
# whole page), so we never pick up shipping icons, app-store badges, etc.
GALLERY_IMG_SELECTORS = (
    ".product__media img, "
    "[id*='MediaGallery'] img, "
    "[id*='ProductMedia'] img, "
    ".product-media img"
)

# A URL only counts as a product image if it lives under this path.
PRODUCT_IMAGE_PATH_MARKER = "/cdn/shop/files/"

# Matches "N results found" / "N result found" in the page <title>,
# which flitit.com sets server-side for every search page.
RESULTS_COUNT_RE = re.compile(r"(\d+)\s*results?\s*found", re.IGNORECASE)

# Matches a trailing "?width=NNN" / "&width=NNN" query param on flitit's
# CDN image URLs, used to detect the same image at different sizes.
WIDTH_PARAM_RE = re.compile(r"[?&]width=(\d+)")


# =========================================================
# HELPERS
# =========================================================
def cell_value_to_str(value) -> str:
    """Convert an Excel cell value to text safely. Numeric-looking cells
    come back from openpyxl as Python floats, e.g. 9555038904369.0 --
    str() on that adds a bogus '.0'. Whole-number floats are converted
    to plain integer text instead."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value)


def clean_search_text(value) -> str:
    """Collapse newlines/tabs/multiple spaces into single spaces and trim."""
    return re.sub(r"\s+", " ", cell_value_to_str(value)).strip()


def build_search_url(query: str) -> str:
    encoded = urllib.parse.quote_plus(clean_search_text(query))
    return SEARCH_URL_TEMPLATE.format(query=encoded)


def to_absolute_url(url: str) -> str:
    if not url:
        return url
    if url.startswith("//"):
        return "https:" + url
    if url.startswith("/"):
        return BASE_URL + url
    return url


def normalize_image_key(url: str) -> str:
    """Strip the width query param so the same image at different sizes
    maps to the same dedupe key."""
    return WIDTH_PARAM_RE.sub("", url)


def image_width(url: str) -> int:
    m = WIDTH_PARAM_RE.search(url)
    return int(m.group(1)) if m else 0


def dedupe_images(urls):
    """De-duplicate by normalized key, keeping the highest-resolution
    variant seen for each unique image, preserving first-seen order."""
    order = []
    best_url = {}
    best_width = {}
    for raw in urls:
        u = to_absolute_url(raw.strip())
        if not u or u.startswith("data:"):
            continue
        key = normalize_image_key(u)
        w = image_width(u)
        if key not in best_url:
            order.append(key)
            best_url[key] = u
            best_width[key] = w
        elif w > best_width[key]:
            best_url[key] = u
            best_width[key] = w
    return [best_url[k] for k in order]


# =========================================================
# SEARCH / SCRAPE CORE
# =========================================================
def get_results_count(page):
    """Read the actual results count from the page <title>
    ("Search: N results found for ..."), which flitit.com renders
    server-side. Returns None if it can't be determined (caller should
    then fall back to a cautious default)."""
    try:
        title = page.title() or ""
    except Exception:
        return None
    m = RESULTS_COUNT_RE.search(title)
    if m:
        return int(m.group(1))
    return None


def search_product(page, search_value: str):
    url = build_search_url(search_value)
    page.goto(url, wait_until="networkidle", timeout=NAV_TIMEOUT)


def get_first_product_link(page):
    """Return the absolute URL of the first real product in the search
    results, or None if the search genuinely returned zero results."""
    count = get_results_count(page)

    # Confirmed zero real results -- do NOT grab any link from the page,
    # since flitit shows unrelated "trending/popular" products in that
    # case, and picking one of those causes wrong-product matches.
    if count == 0:
        return None

    try:
        links = page.locator(FIRST_PRODUCT_LINK_SELECTOR)
        if links.count() == 0:
            return None
        href = links.first.get_attribute("href")
        if not href:
            return None
        return to_absolute_url(href)
    except Exception:
        return None


def extract_description(page) -> str:
    try:
        page.wait_for_selector(DESCRIPTION_SELECTOR, timeout=10000)
        loc = page.locator(DESCRIPTION_SELECTOR).first
        text = loc.inner_text()
        return text.strip() if text else ""
    except PlaywrightTimeoutError:
        return ""
    except Exception:
        return ""


def extract_images(page) -> str:
    candidates = []
    try:
        img_locators = page.locator(GALLERY_IMG_SELECTORS)
        count = img_locators.count()
        for i in range(count):
            img = img_locators.nth(i)
            src = (
                img.get_attribute("src")
                or img.get_attribute("data-src")
                or img.get_attribute("srcset")
                or img.get_attribute("data-srcset")
            )
            if not src:
                continue
            # srcset can hold multiple "url widthDescriptor" entries.
            if "," in src and " " in src:
                src = src.split(",")[0].strip().split(" ")[0].strip()
            if PRODUCT_IMAGE_PATH_MARKER not in src:
                continue
            candidates.append(src)
    except Exception:
        pass

    return " ; ".join(dedupe_images(candidates))


def process_product(page, barcode, item_name, row_num: int, total: int):
    """barcode is accepted for signature compatibility but is expected to
    always be None now that barcode matching is disabled -- only
    item_name is used to search/match."""
    result = {"Description": "", "Image URLs": "", "Status": "NO_RESULT"}

    try:
        product_link = None
        item_name_str = clean_search_text(item_name) if item_name not in (None, "") else ""

        if item_name_str:
            print(f"[{row_num}/{total}] Searching Name...")
            search_product(page, item_name_str)
            product_link = get_first_product_link(page)

        if not product_link:
            print(f"[{row_num}/{total}] NO_RESULT")
            return result

        page.goto(product_link, wait_until="networkidle", timeout=NAV_TIMEOUT)

        result["Description"] = extract_description(page)
        result["Image URLs"] = extract_images(page)
        result["Status"] = "MATCHED"
        print(f"[{row_num}/{total}] MATCHED")

    except Exception as exc:
        error_msg = f"ERROR: {exc}"
        print(f"[{row_num}/{total}] {error_msg}")
        result["Status"] = error_msg

    return result


# =========================================================
# EXCEL (openpyxl-only, preserves embedded images/formatting)
# =========================================================
def find_header_row(ws, required, max_scan=10):
    required_lower = [r.strip().lower() for r in required]
    for r in range(1, min(max_scan, ws.max_row) + 1):
        row_values = {
            str(ws.cell(row=r, column=c).value).strip().lower()
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value is not None
        }
        if all(req in row_values for req in required_lower):
            return r
    return 1


def find_column(ws, header_row, name):
    target = name.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip().lower() == target:
            return c
    return None


def find_real_last_data_row(ws, header_row, barcode_col, item_name_col):
    """ws.max_row can be inflated by leftover cell formatting (borders,
    fill color, etc.) on rows that have no actual data -- Excel/openpyxl
    still counts those as "used" rows. Walk up from the bottom until we
    hit a row that actually has an Item Name value, and treat that as
    the real end of the data. (barcode_col kept in the signature for
    backward compatibility but is not used.)"""
    for r in range(ws.max_row, header_row, -1):
        n = ws.cell(row=r, column=item_name_col).value
        if n is not None and str(n).strip() != "":
            return r
    return header_row


def prepare_master_workbook(input_file: str, row_limit):
    """Load the original file, append the 3 new (empty) columns, physically
    delete trailing phantom-empty rows (rows with leftover formatting but
    no Item Name), and optionally truncate further to row_limit data rows
    (tail-only deletion, so it never disturbs the rows/images we're
    keeping). row_limit=None keeps every real data row."""
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    header_row = find_header_row(ws, [ITEM_NAME_HEADER])
    barcode_col = None
    item_name_col = find_column(ws, header_row, ITEM_NAME_HEADER)

    if item_name_col is None:
        raise ValueError(
            f"Couldn't find '{ITEM_NAME_HEADER}' column in header row {header_row}."
        )

    last_col = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip() != "":
            last_col = c

    new_col_idx = {}
    for offset, name in enumerate(NEW_COLUMNS, start=1):
        col = last_col + offset
        ws.cell(row=header_row, column=col, value=name)
        new_col_idx[name] = col

    # Drop trailing phantom-empty rows (no Item Name) so we don't loop
    # over hundreds of blank formatted rows and so the output files
    # don't carry them either.
    real_last_row = find_real_last_data_row(ws, header_row, barcode_col, item_name_col)
    if ws.max_row > real_last_row:
        ws.delete_rows(real_last_row + 1, ws.max_row - real_last_row)

    last_data_row = real_last_row
    if row_limit is not None:
        wanted_last_row = header_row + row_limit
        if last_data_row > wanted_last_row:
            ws.delete_rows(wanted_last_row + 1, last_data_row - wanted_last_row)
        last_data_row = min(last_data_row, wanted_last_row)

    return wb, ws, header_row, last_data_row, barcode_col, item_name_col, new_col_idx


def write_found_notfound(master_path: str, header_row: int, last_data_row: int, status_col: int):
    base, ext = os.path.splitext(master_path)
    found_path = f"{base}_found{ext}"
    notfound_path = f"{base}_notfound{ext}"

    for out_path, keep_matched in ((found_path, True), (notfound_path, False)):
        shutil.copy(master_path, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        for r in range(header_row + 1, last_data_row + 1):
            status_val = ws.cell(row=r, column=status_col).value or ""
            is_matched = str(status_val) == "MATCHED"
            hide = (is_matched != keep_matched)
            ws.row_dimensions[r].hidden = hide
        wb.save(out_path)

    return found_path, notfound_path


# =========================================================
# MAIN
# =========================================================
def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE

    row_limit = ROW_LIMIT
    if len(sys.argv) > 2:
        arg = sys.argv[2].strip().lower()
        row_limit = None if arg in ("all", "0", "none") else int(arg)

    if not os.path.isfile(input_file):
        print(f"الملف مش موجود: {input_file}")
        return 1

    print(f"Input file : {input_file}")
    print(f"Row limit  : {row_limit if row_limit is not None else 'ALL'}\n")

    wb, ws, header_row, last_data_row, barcode_col, item_name_col, new_col_idx = (
        prepare_master_workbook(input_file, row_limit)
    )
    desc_col = new_col_idx["Description"]
    img_col = new_col_idx["Image URLs"]
    status_col = new_col_idx["Status"]

    total = last_data_row - header_row

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.set_default_timeout(NAV_TIMEOUT)

        for i, r in enumerate(range(header_row + 1, last_data_row + 1), start=1):
            item_name = ws.cell(row=r, column=item_name_col).value

            item_name_empty = cell_value_to_str(item_name).strip() == ""

            if item_name_empty:
                continue  # leave the new columns blank for this row

            outcome = process_product(page, None, item_name, i, total)

            ws.cell(row=r, column=desc_col, value=outcome["Description"])
            ws.cell(row=r, column=img_col, value=outcome["Image URLs"])
            ws.cell(row=r, column=status_col, value=outcome["Status"])

        browser.close()

    base, ext = os.path.splitext(input_file)
    master_filename = os.path.basename(base) + f"_scraped{ext}"
    master_path = os.path.join(os.getcwd(), master_filename)
    wb.save(master_path)
    print(f"\nMaster file saved -> {master_path}")

    found_path, notfound_path = write_found_notfound(
        master_path, header_row, last_data_row, status_col
    )
    print(f"Found file    -> {found_path}")
    print(f"Not-found file -> {notfound_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())