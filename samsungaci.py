#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SamsungACI WooCommerce Scraper
--------------------------------
Reads SKUs from an Excel file, searches on samsungaci.com, opens the first
product result, and extracts the product description (HTML) and image URLs.

Output: an Excel file with the input columns plus:
  - Description: clean product description (HTML tags removed)
  - Image Src:  semicolon-separated absolute image URLs
  - Source_URL: product page URL clicked
  - Found:       YES/NO

Usage (PowerShell / CMD):

  python samsungaci_scraper.py \
    --in "input.xlsx" \
    --out "output.xlsx" \
    --sku-col "Variant SKU" \
    --sheet "Sheet1" \
    --header-row 0 \
    --pause 1.0 \
    --headful

  # Or append to original file:
  python samsungaci_scraper.py \
    --in "Data/HA&AV.xlsx" \
    --sku-col "Item" \
    --sheet "HA" \
    --header-row 3 \
    --append

Notes:
- Set --headful to see the browser. Omit for headless.
- If a SKU yields no product results, the row is left blank and Found=NO.
- Use --append to update the original file instead of creating a new one.
- Requires: pandas, selenium, openpyxl (for .xlsx IO)

pip install -U pandas selenium openpyxl

Test quickly on a few SKUs using --sample 5 to process just the first 5 rows.
"""

from __future__ import annotations

import argparse
import sys
import time
import re
import urllib.parse
from dataclasses import dataclass
from typing import List, Optional, Tuple, Iterable, Set

import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException


SEARCH_TEMPLATE = (
    "https://samsungaci.com/?s={query}&post_type=product&type_aws=true&lang=en"
)


@dataclass
class ScrapeResult:
    body_html: str = ""
    image_urls: List[str] = None
    source_url: str = ""
    found: bool = False


# ------------------------------ Selenium helpers ------------------------------

def build_driver(headful: bool = False, user_data_dir: Optional[str] = None) -> webdriver.Chrome:
    opts = webdriver.ChromeOptions()
    if not headful:
        # modern headless (Chrome >= 109)
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1366,900")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    if user_data_dir:
        opts.add_argument(f"--user-data-dir={user_data_dir}")

    # Let Selenium Manager resolve chromedriver automatically
    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(45)
    return driver


def first_present(driver: webdriver.Chrome, selectors: Iterable[Tuple[By, str]], timeout: float = 10):
    """Return the first element that appears among a list of selectors, else None."""
    end = time.time() + timeout
    while time.time() < end:
        for by, sel in selectors:
            try:
                el = driver.find_element(by, sel)
                if el:
                    return el
            except NoSuchElementException:
                continue
        time.sleep(0.2)
    return None


def all_present(driver: webdriver.Chrome, selectors: Iterable[Tuple[By, str]]) -> List[Tuple[By, str, List]]:
    out = []
    for by, sel in selectors:
        try:
            els = driver.find_elements(by, sel)
            if els:
                out.append((by, sel, els))
        except Exception:
            pass
    return out


# ------------------------------ Core scraping ------------------------------

def open_first_product(driver: webdriver.Chrome, sku: str, pause: float = 0.75) -> Optional[str]:
    """Search for SKU and click the first product. Return product URL or None."""
    q = urllib.parse.quote_plus(str(sku).strip())
    search_url = SEARCH_TEMPLATE.format(query=q)
    driver.get(search_url)

    # Typical WooCommerce search result product link selectors (robust list)
    candidate_links = [
        (By.CSS_SELECTOR, "ul.products li.product a.woocommerce-LoopProduct-link"),
        (By.CSS_SELECTOR, "ul.products li.product a[href*='/product/']"),
        (By.CSS_SELECTOR, "li.product a[href*='/product/']"),
        (By.CSS_SELECTOR, "a.woocommerce-LoopProduct-link"),
        (By.CSS_SELECTOR, "a[href*='/product/']"),
    ]

    link_el = first_present(driver, candidate_links, timeout=12)
    if not link_el:
        # Try to detect an explicit "no results" message (optional)
        return None

    url = link_el.get_attribute("href") or ""
    if not url:
        try:
            link_el.click()
            time.sleep(pause)
            url = driver.current_url
        except Exception:
            return None
    else:
        driver.get(url)

    # Wait for product page structure
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.product, div.woocommerce"))
        )
        time.sleep(pause)
    except TimeoutException:
        return None

    return driver.current_url


def clean_html_text(html_content: str) -> str:
    """Remove HTML tags and clean up text content."""
    if not html_content:
        return ""
    
    from bs4 import BeautifulSoup
    import re
    
    try:
        # Parse HTML
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()
        
        # Get text content
        text = soup.get_text()
        
        # Clean up whitespace
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = ' '.join(chunk for chunk in chunks if chunk)
        
        # Remove extra spaces
        text = re.sub(r'\s+', ' ', text)
        
        return text.strip()
    except Exception:
        # Fallback: simple regex-based HTML removal
        import re
        # Remove HTML tags
        text = re.sub(r'<[^>]+>', '', html_content)
        # Clean up whitespace
        text = re.sub(r'\s+', ' ', text)
        return text.strip()


def extract_description(driver: webdriver.Chrome) -> str:
    """Return clean text description (HTML tags removed)."""
    # Long/Full description tab content usually
    desc_selectors = [
        (By.CSS_SELECTOR, "div.woocommerce-Tabs-panel--description"),
        (By.CSS_SELECTOR, "div#tab-description"),
        (By.CSS_SELECTOR, "div.product div[itemprop='description']"),
        # Short description fallback
        (By.CSS_SELECTOR, "div.woocommerce-product-details__short-description"),
    ]

    el = first_present(driver, desc_selectors, timeout=6)
    if el is None:
        return ""

    # Try to get HTML content first
    html = el.get_attribute("innerHTML") or ""
    if html.strip():
        return clean_html_text(html)

    # Fallback to text content
    txt = el.text.strip()
    return txt if txt else ""


def is_valid_image_url(url: str) -> bool:
    """Check if URL is a valid image URL."""
    if not url or not isinstance(url, str):
        return False
    return bool(re.search(r"\.(?:jpg|jpeg|png|webp|gif)(?:[?#].*)?$", url, re.I))


def extract_images(driver: webdriver.Chrome) -> List[str]:
    """Return an ordered, deduplicated list of full-sized image URLs."""
    image_candidates: List[Tuple[str, int]] = []  # (url, priority_score)

    # Common Woo gallery containers
    gallery_containers = [
        (By.CSS_SELECTOR, "div.woocommerce-product-gallery"),
        (By.CSS_SELECTOR, "div.woocommerce-product-gallery__wrapper"),
        (By.CSS_SELECTOR, "div.product div.images"),
        (By.CSS_SELECTOR, "div.product-gallery, div.images"),
    ]

    # Try to gather from gallery containers first (highest priority)
    found_any = all_present(driver, gallery_containers)
    if found_any:
        for _, _, containers in found_any:
            for c in containers:
                # 1) Anchors wrapping images (usually full-size)
                for a in c.find_elements(By.CSS_SELECTOR, "a[href]"):
                    href = (a.get_attribute("href") or "").strip()
                    if is_valid_image_url(href):
                        # Check for size indicators in URL
                        priority = get_image_priority(href)
                        image_candidates.append((href, priority))
                
                # 2) Data attributes for large images
                for img in c.find_elements(By.CSS_SELECTOR, "img"):
                    for attr, priority in [("data-large_image", 100), ("data-src", 80), ("data-full", 90), ("src", 60)]:
                        val = (img.get_attribute(attr) or "").strip()
                        if is_valid_image_url(val):
                            # Boost priority for gallery images
                            final_priority = priority + get_image_priority(val)
                            image_candidates.append((val, final_priority))

    # Fallback: whole page scan (lower priority)
    if not image_candidates:
        for img in driver.find_elements(By.CSS_SELECTOR, "img[src]"):
            src = (img.get_attribute("src") or "").strip()
            if is_valid_image_url(src):
                priority = get_image_priority(src)
                image_candidates.append((src, priority))

    # Remove duplicates using normalized URLs and sort by priority (highest first)
    seen_urls: Set[str] = set()  # Original URLs
    seen_normalized: Set[str] = set()  # Normalized URLs for duplicate detection
    unique_images: List[Tuple[str, int]] = []
    
    for url, priority in image_candidates:
        normalized = normalize_image_url(url)
        
        # Skip if we've already seen this normalized URL (duplicate)
        if normalized in seen_normalized:
            continue
            
        # Skip if we've already seen this exact URL
        if url in seen_urls:
            continue
            
        seen_urls.add(url)
        seen_normalized.add(normalized)
        unique_images.append((url, priority))
    
    # Sort by priority (highest first) and return URLs
    unique_images.sort(key=lambda x: x[1], reverse=True)
    return [url for url, _ in unique_images]


def normalize_image_url(url: str) -> str:
    """Normalize image URL to help detect duplicates."""
    if not url:
        return ""
    
    # Remove common query parameters that don't affect the image
    url = url.split('?')[0].split('#')[0]
    
    # Remove size suffixes to get base image name
    import re
    # Remove common size patterns like -1200x1200, -300x300, etc.
    url = re.sub(r'-\d+x\d+', '', url)
    # Remove other size indicators
    url = re.sub(r'-(?:small|medium|large|thumb|mini|scaled|resized)', '', url, flags=re.I)
    
    return url.lower().strip()


def get_image_priority(url: str) -> int:
    """Calculate priority score for image URL based on size indicators."""
    if not url:
        return 0
    
    url_lower = url.lower()
    priority = 0
    
    # Size indicators in URL
    if any(size in url_lower for size in ['-150x150', '-300x300', '-600x600', '-768x768']):
        priority -= 20  # Thumbnail sizes
    if any(size in url_lower for size in ['-1024x1024', '-1200x1200', '-2048x2048']):
        priority += 30  # Large sizes
    if any(size in url_lower for size in ['-scaled', '-resized', 'thumbnail']):
        priority -= 10  # Scaled/resized images
    if any(size in url_lower for size in ['full', 'original', 'large']):
        priority += 50  # Full/original size indicators
    
    # File size indicators in URL
    if any(size in url_lower for size in ['_large', '_big', '_full', '_original']):
        priority += 40
    if any(size in url_lower for size in ['_small', '_thumb', '_mini']):
        priority -= 30
    
    # Gallery-specific indicators
    if 'gallery' in url_lower or 'product' in url_lower:
        priority += 20
    
    return priority


def scrape_with_retry(driver: webdriver.Chrome, sku: str, pause: float = 0.75, max_retries: int = 3) -> ScrapeResult:
    """Scrape with retry mechanism for better reliability."""
    for attempt in range(max_retries):
        try:
            return scrape_one(driver, sku, pause)
        except Exception as e:
            if attempt == max_retries - 1:
                raise e
            print(f"    Retry {attempt + 1}/{max_retries - 1} for {sku}: {e}")
            time.sleep(2 ** attempt)  # Exponential backoff
    return ScrapeResult()


def scrape_one(driver: webdriver.Chrome, sku: str, pause: float = 0.75) -> ScrapeResult:
    res = ScrapeResult(body_html="", image_urls=[], source_url="", found=False)

    product_url = open_first_product(driver, sku, pause=pause)
    if not product_url:
        return res

    res.source_url = product_url
    res.body_html = extract_description(driver)
    res.image_urls = extract_images(driver)
    res.found = bool(res.body_html or res.image_urls)
    return res


# ------------------------------ I/O & main ------------------------------

def read_excel(path: str, sheet: Optional[str] = None, header_row: int = 0) -> pd.DataFrame:
    try:
        if sheet:
            return pd.read_excel(path, sheet_name=sheet, header=header_row)
        return pd.read_excel(path, header=header_row)
    except Exception as e:
        raise SystemExit(f"Failed to read Excel: {path}\n{e}")


def ensure_output_cols(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["Description", "Image Src", "Source_URL", "Found"]:
        if col not in df.columns:
            df[col] = ""
        # Ensure proper data types to avoid pandas warnings
        df[col] = df[col].astype(str)
    return df


def main():
    ap = argparse.ArgumentParser(description="Scrape samsungaci.com product description & images by SKU from Excel.")
    ap.add_argument("--in", dest="inp", required=True, help="Input Excel .xlsx path")
    ap.add_argument("--out", dest="out", required=False, help="Output Excel .xlsx path (not needed with --append)")
    ap.add_argument("--sku-col", dest="sku_col", default="Variant SKU", help="Column containing SKUs (default: Variant SKU)")
    ap.add_argument("--sheet", dest="sheet", default=None, help="Optional Excel sheet name")
    ap.add_argument("--header-row", dest="header_row", type=int, default=0, help="Zero-based header row index (default: 0)")
    ap.add_argument("--pause", dest="pause", type=float, default=0.75, help="Small wait after page loads (sec)")
    ap.add_argument("--headful", dest="headful", action="store_true", help="Show browser window")
    ap.add_argument("--profile", dest="profile", default=None, help="Optional Chrome user-data-dir for logged-in session")
    ap.add_argument("--sample", dest="sample", type=int, default=None, help="Process only the first N rows (debug)")
    ap.add_argument("--retries", dest="retries", type=int, default=3, help="Max retry attempts for failed requests (default: 3)")
    ap.add_argument("--append", dest="append", action="store_true", help="Append data to original file instead of creating new file")

    args = ap.parse_args()

    # Validate arguments
    if not args.append and not args.out:
        raise SystemExit("Error: Either --out or --append must be specified")
    if args.append and args.out:
        raise SystemExit("Error: Cannot use both --out and --append. Choose one.")

    df = read_excel(args.inp, sheet=args.sheet, header_row=args.header_row)
    if args.sku_col not in df.columns:
        raise SystemExit(f"SKU column '{args.sku_col}' not found in input file. Columns: {list(df.columns)}")

    # Validate SKU column has data
    if df[args.sku_col].isna().all():
        raise SystemExit(f"No valid SKUs found in column '{args.sku_col}'. All values are empty/NaN.")
    
    # Count valid SKUs
    valid_skus = df[args.sku_col].notna().sum()
    print(f"Found {valid_skus} valid SKUs to process")

    if args.sample:
        df = df.head(args.sample).copy()
        print(f"Processing sample of {len(df)} rows")

    df = ensure_output_cols(df)

    driver = build_driver(headful=args.headful, user_data_dir=args.profile)

    try:
        total_rows = len(df)
        processed = 0
        found_count = 0
        
        for idx, row in df.iterrows():
            sku_raw = row[args.sku_col]
            sku = "" if pd.isna(sku_raw) else str(sku_raw).strip()
            
            if not sku:
                df.at[idx, "Found"] = "NO"
                processed += 1
                continue

            print(f"[{processed + 1}/{total_rows}] Processing SKU: {sku}")
            
            try:
                result = scrape_with_retry(driver, sku, pause=args.pause, max_retries=args.retries)
                df.at[idx, "Description"] = result.body_html
                df.at[idx, "Image Src"] = ";".join(result.image_urls or [])
                df.at[idx, "Source_URL"] = result.source_url
                df.at[idx, "Found"] = "YES" if result.found else "NO"
                
                if result.found:
                    found_count += 1
                    print(f"  ✓ Found: {len(result.image_urls)} images, {len(result.body_html)} chars description")
                    # Show first few image URLs for verification
                    if result.image_urls:
                        print(f"    Sample images:")
                        for i, img_url in enumerate(result.image_urls[:3]):
                            print(f"      {i+1}. {img_url}")
                        if len(result.image_urls) > 3:
                            print(f"      ... and {len(result.image_urls) - 3} more")
                else:
                    print(f"  ✗ Not found")
                    
            except Exception as e:
                df.at[idx, "Found"] = "NO"
                print(f"  ✗ Error: {e}")
            
            processed += 1
            
        print(f"\nCompleted: {found_count}/{processed} SKUs found")

    finally:
        driver.quit()

    try:
        if args.append:
            # Append to original file, preserving format
            from openpyxl import load_workbook
            
            # Load the original workbook to preserve formatting
            wb = load_workbook(args.inp)
            sheet_name = args.sheet or 'Sheet1'
            
            # Get the original sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Find the header row
            header_row = args.header_row + 1  # Convert to 1-based indexing
            
            # Add new columns if they don't exist
            new_columns = ["Description", "Image Src", "Source_URL", "Found"]
            col_idx = ws.max_column + 1
            
            for col_name in new_columns:
                # Check if column already exists
                col_exists = False
                for cell in ws[header_row]:
                    if cell.value == col_name:
                        col_exists = True
                        break
                
                if not col_exists:
                    # Add column header
                    ws.cell(row=header_row, column=col_idx, value=col_name)
                    col_idx += 1
            
            # Find the column indices for new columns
            desc_col = None
            img_col = None
            url_col = None
            found_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value == "Description":
                    desc_col = col
                elif cell_value == "Image Src":
                    img_col = col
                elif cell_value == "Source_URL":
                    url_col = col
                elif cell_value == "Found":
                    found_col = col
            
            # Update data rows
            for idx, row in df.iterrows():
                excel_row = idx + header_row + 1  # Convert to Excel row number
                
                if desc_col and not pd.isna(row.get("Description", "")):
                    ws.cell(row=excel_row, column=desc_col, value=row.get("Description", ""))
                if img_col and not pd.isna(row.get("Image Src", "")):
                    ws.cell(row=excel_row, column=img_col, value=row.get("Image Src", ""))
                if url_col and not pd.isna(row.get("Source_URL", "")):
                    ws.cell(row=excel_row, column=url_col, value=row.get("Source_URL", ""))
                if found_col and not pd.isna(row.get("Found", "")):
                    ws.cell(row=excel_row, column=found_col, value=row.get("Found", ""))
            
            # Save the workbook
            wb.save(args.inp)
            print(f"Updated → {args.inp} (sheet: {sheet_name})")
        else:
            # Create new file
            df.to_excel(args.out, index=False)
            print(f"Saved → {args.out}")
    except Exception as e:
        if args.append:
            raise SystemExit(f"Failed updating Excel: {args.inp}\n{e}")
        else:
            raise SystemExit(f"Failed writing Excel: {args.out}\n{e}")


if __name__ == "__main__":
    main()
