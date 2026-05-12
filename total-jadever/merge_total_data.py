#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Merge upload_template_final.xlsx data into the original file based on BAR CODE.
Preserves images and formatting in the original file using openpyxl.

Note: openpyxl has limited support for images. If images are lost, consider:
1. Using xlwings (requires Excel installed) for better image preservation
2. Manually copying images after merging
3. Using a different approach that preserves images better
"""

import sys
import io
if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil

# Get the directory where this script is located
SCRIPT_DIR = Path(__file__).parent.absolute()

# File paths
ORIGINAL_FILE = Path(r"D:\JafarShop\Scrapers\total\jadever\Jadever 6M - Copy.xlsx")
TEMPLATE_FILE = Path(r"D:\JafarShop\Scrapers\total\jadever\upload_template_final.xlsx")
OUTPUT_FILE = Path(r"D:\JafarShop\Scrapers\total\jadever\Jadever 6M - Copy.xlsx")  # Overwrite original
BACKUP_FILE = Path(r"D:\JafarShop\Scrapers\total\jadever\Jadever 6M - Copy_backup.xlsx")  # Backup before merging

def normalize_item_no(value):
    """Normalize BAR CODE for matching - handles numeric, string, leading zeros, etc."""
    if value is None or pd.isna(value):
        return None
    
    # Convert to string and strip whitespace
    value_str = str(value).strip()
    
    # Handle NaN string representation
    if value_str.lower() in ['nan', 'none', '']:
        return None
    
    # Try to convert to number to remove leading zeros and normalize
    try:
        # If it's a number, convert to int then back to string to remove leading zeros
        num_value = float(value_str)
        if num_value == int(num_value):
            return str(int(num_value))
        else:
            return str(num_value)
    except (ValueError, TypeError):
        # If it's not a number, return uppercase for case-insensitive matching
        return value_str.upper()

def find_barcode_column(df):
    """Find the BAR CODE column in the dataframe. Only looks for BAR CODE, no fallbacks."""
    for col in df.columns:
        col_str = str(col).strip()
        # Exact matches for BAR CODE variations
        if col_str in ["BAR CODE", "BARCODE", "Barcode", "barcode", "Bar Code", "باركود"]:
            return col
        # Contains "bar" and "code" (case insensitive)
        if "bar" in col_str.lower() and "code" in col_str.lower():
            return col
    
    return None

def find_column_index(ws, column_name):
    """Find the column index (1-based) for a given column name in the first row."""
    for cell in ws[1]:
        if str(cell.value).strip() == str(column_name).strip():
            return cell.column
    return None

def main():
    """Main function to merge the data while preserving images."""
    print("Creating backup of original file...")
    try:
        shutil.copy2(ORIGINAL_FILE, BACKUP_FILE)
        print(f"Backup created: {BACKUP_FILE}")
    except Exception as e:
        print(f"Warning: Could not create backup: {e}")
    
    print("\nReading template file...")
    try:
        df_template = pd.read_excel(TEMPLATE_FILE)
        print(f"Template file shape: {df_template.shape}")
        print(f"Template columns: {list(df_template.columns)}")
        print(f"First few rows of template:")
        print(df_template.head())
    except Exception as e:
        print(f"Error reading template file: {e}")
        return
    
    # Find BAR CODE column in template
    print(f"\nLooking for BAR CODE column in template file...")
    print(f"Available columns: {list(df_template.columns)}")
    template_barcode_col = find_barcode_column(df_template)
    if not template_barcode_col:
        print(f"ERROR: Could not find 'BAR CODE' column in template file.")
        print(f"Available columns: {list(df_template.columns)}")
        return
    
    print(f"✓ Found BAR CODE column: '{template_barcode_col}'")
    print(f"Sample BAR CODE values from template:")
    print(df_template[template_barcode_col].head(10).tolist())
    
    # Get columns from template that are not BAR CODE (to append)
    columns_to_append = [col for col in df_template.columns if col != template_barcode_col]
    print(f"\nColumns to append from template: {columns_to_append}")
    
    # Create a mapping: BAR CODE -> dict of values to append (with normalized keys)
    # Also keep original values for fallback matching
    template_dict = {}
    template_dict_original = {}  # For exact string matching fallback
    template_original_values = {}  # Keep original values for debugging
    
    print(f"\nProcessing template barcodes...")
    for idx, row in df_template.iterrows():
        barcode_raw = row[template_barcode_col]
        barcode_normalized = normalize_item_no(barcode_raw)
        barcode_str = str(barcode_raw).strip() if barcode_raw is not None else ""
        
        # Store all possible variations for maximum matching
        if barcode_normalized:
            template_dict[barcode_normalized] = {col: row[col] for col in columns_to_append}
            template_original_values[barcode_normalized] = barcode_raw
        
        # Store with original string format (multiple variations)
        if barcode_str and barcode_str.lower() not in ['nan', 'none', '']:
            # Uppercase version
            template_dict_original[barcode_str.upper()] = {col: row[col] for col in columns_to_append}
            # Original case
            template_dict_original[barcode_str] = {col: row[col] for col in columns_to_append}
            # Lowercase version
            template_dict_original[barcode_str.lower()] = {col: row[col] for col in columns_to_append}
            
            # Also try without any normalization (raw string)
            if barcode_raw is not None:
                barcode_raw_str = str(barcode_raw)
                template_dict_original[barcode_raw_str] = {col: row[col] for col in columns_to_append}
                template_dict_original[barcode_raw_str.upper()] = {col: row[col] for col in columns_to_append}
                template_dict_original[barcode_raw_str.lower()] = {col: row[col] for col in columns_to_append}
        
        # Debug first few
        if idx < 3:
            print(f"  Template row {idx}: raw='{barcode_raw}' (type: {type(barcode_raw)}), normalized='{barcode_normalized}', str='{barcode_str}'")
    
    print(f"Created mapping for {len(template_dict)} items from template")
    print(f"Sample template BAR CODEs: {list(template_original_values.values())[:5]}")
    print(f"Sample normalized template keys: {list(template_dict.keys())[:10]}")
    print(f"Sample template_dict_original keys: {list(template_dict_original.keys())[:10]}")
    
    # Debug: Check if TFCLI42021 is in template
    debug_item = "TFCLI42021"
    debug_found = False
    for key in template_dict.keys():
        if debug_item.upper() in str(key).upper():
            print(f"\nDEBUG: Found '{debug_item}' in template_dict as key: '{key}'")
            debug_found = True
    for key in template_dict_original.keys():
        if debug_item.upper() in str(key).upper():
            print(f"DEBUG: Found '{debug_item}' in template_dict_original as key: '{key}'")
            debug_found = True
    if not debug_found:
        print(f"\nDEBUG: '{debug_item}' NOT found in template dictionaries")
        print(f"Sample template keys: {list(template_dict.keys())[:10]}")
    
    # Load the original workbook (preserves images and formatting)
    # Note: openpyxl has limited image support - images may not be fully preserved
    print("\nLoading original workbook (preserving images)...")
    print("⚠️  Note: openpyxl has limited image support. Images may not be fully preserved.")
    
    try:
        # Try loading without keep_vba first (most files don't have VBA)
        wb = load_workbook(ORIGINAL_FILE, data_only=False, keep_links=False)
        ws = wb.active
        print(f"✓ Original workbook loaded. Active sheet: {ws.title}")
        print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
        print(f"  Workbook has {len(wb.worksheets)} worksheet(s)")
        
        # Check if workbook has images (openpyxl can detect some image types)
        try:
            if hasattr(ws, '_images') and len(ws._images) > 0:
                print(f"  ✓ Detected {len(ws._images)} image(s) in the worksheet")
            else:
                print(f"  ⚠️  No images detected by openpyxl (may still exist in file)")
        except:
            pass
            
    except Exception as e:
        print(f"Error loading original workbook: {e}")
        print("Trying with different options...")
        try:
            # Try with minimal options
            wb = load_workbook(ORIGINAL_FILE)
            ws = wb.active
            print(f"✓ Original workbook loaded (basic mode). Active sheet: {ws.title}")
        except Exception as e2:
            print(f"❌ Error loading original workbook: {e2}")
            import traceback
            traceback.print_exc()
            return
    
    # Read the original file with pandas to find the BAR CODE column
    print("\nReading original file structure...")
    try:
        df_original = pd.read_excel(ORIGINAL_FILE)
        print(f"Original file shape: {df_original.shape}")
        print(f"Original columns: {list(df_original.columns)}")
        print(f"First few rows of original:")
        print(df_original.head())
    except Exception as e:
        print(f"Error reading original file: {e}")
        return
    
    # Find BAR CODE column in original
    print(f"\nLooking for BAR CODE column in original file...")
    print(f"Available columns: {list(df_original.columns)}")
    original_barcode_col = find_barcode_column(df_original)
    if not original_barcode_col:
        print(f"ERROR: Could not find 'BAR CODE' column in original file.")
        print(f"Available columns: {list(df_original.columns)}")
        return
    
    print(f"✓ Found BAR CODE column: '{original_barcode_col}'")
    print(f"Sample BAR CODE values from original:")
    print(df_original[original_barcode_col].head(10).tolist())
    
    # Verify we're using BAR CODE
    if "bar" in original_barcode_col.lower() and "code" in original_barcode_col.lower():
        print(f"✓ Confirmed: Using BAR CODE column for matching")
    else:
        print(f"⚠️  WARNING: Column name '{original_barcode_col}' doesn't look like BAR CODE!")
    
    # Find the column index for BAR CODE in the worksheet
    print(f"\nFinding '{original_barcode_col}' column in worksheet...")
    barcode_col_idx = find_column_index(ws, original_barcode_col)
    if not barcode_col_idx:
        print(f"ERROR: Could not find '{original_barcode_col}' column in worksheet")
        print(f"First row values: {[cell.value for cell in ws[1]]}")
        return
    
    print(f"✓ BAR CODE column found at column {get_column_letter(barcode_col_idx)} (index {barcode_col_idx})")
    
    # Find the last column with data/header
    # First, check what's in row 1 to preserve existing headers
    print("\nChecking existing headers in row 1...")
    existing_headers = []
    max_header_col = 0
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col_idx).value
        if header_value:
            existing_headers.append(str(header_value).strip())
            max_header_col = max(max_header_col, col_idx)
        else:
            existing_headers.append("")
    
    print(f"  Found {len([h for h in existing_headers if h])} existing header(s)")
    print(f"  Last column with header: {get_column_letter(max_header_col)} (index {max_header_col})")
    if existing_headers:
        print(f"  Sample headers: {[h for h in existing_headers[:10] if h]}")
    
    # Find the last column with a header or data - use the maximum
    last_col = max(ws.max_column, max_header_col)
    print(f"  Will add new columns starting from column {get_column_letter(last_col + 1)}")
    
    # Add new column headers (only if they don't already exist)
    print("\nAdding new column headers...")
    headers_added = 0
    new_col_start = last_col + 1
    col_offset = 0
    
    for col_name in columns_to_append:
        # Check if this column name already exists in headers
        if col_name not in existing_headers:
            new_col = new_col_start + col_offset
            # Make sure we're not overwriting anything
            existing_value = ws.cell(row=1, column=new_col).value
            if existing_value:
                print(f"  ⚠️  Column {get_column_letter(new_col)} already has value '{existing_value}', skipping...")
                col_offset += 1
                new_col = new_col_start + col_offset
            
            ws.cell(row=1, column=new_col, value=col_name)
            print(f"  ✓ Added column '{col_name}' at column {get_column_letter(new_col)}")
            headers_added += 1
            col_offset += 1
        else:
            print(f"  ⊘ Skipped column '{col_name}' - already exists in row 1")
    
    if headers_added == 0:
        print("  ⚠️  No new headers were added (all columns already exist)")
    else:
        print(f"  ✓ Added {headers_added} new column header(s)")
    
    # Create a mapping of column names to their column indices for data filling
    column_name_to_index = {}
    for col_name in columns_to_append:
        # Find which column index this header was added to
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == col_name:
                column_name_to_index[col_name] = col_idx
                break
    
    print(f"\nColumn mapping for data filling:")
    for col_name, col_idx in column_name_to_index.items():
        print(f"  '{col_name}' -> column {get_column_letter(col_idx)}")
    
    # Fill in the data by matching BAR CODE
    print("\nMatching and filling data...")
    print(f"Total rows to process: {ws.max_row - 1}")
    matched_count = 0
    unmatched_items = []
    matched_items = []
    
    # Debug: Show first few barcodes from worksheet
    print("\nFirst 5 BAR CODE values from worksheet:")
    for row_idx in range(2, min(7, ws.max_row + 1)):
        barcode_cell = ws.cell(row=row_idx, column=barcode_col_idx)
        barcode_val = barcode_cell.value
        barcode_norm = normalize_item_no(barcode_val)
        barcode_str = str(barcode_val).strip() if barcode_val is not None else ""
        print(f"  Row {row_idx}: raw='{barcode_val}' (type: {type(barcode_val)}), normalized='{barcode_norm}', str='{barcode_str}'")
    
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        barcode_cell = ws.cell(row=row_idx, column=barcode_col_idx)
        barcode_raw = barcode_cell.value
        barcode_normalized = normalize_item_no(barcode_raw)
        barcode_str = str(barcode_raw).strip() if barcode_raw is not None else ""
        barcode_str_upper = barcode_str.upper() if barcode_str else ""
        
        matched_data = None
        
        # Try all possible matching strategies
        # 1. Normalized matching
        if barcode_normalized and barcode_normalized in template_dict:
            matched_data = template_dict[barcode_normalized]
        # 2. Uppercase string matching
        elif barcode_str_upper and barcode_str_upper in template_dict_original:
            matched_data = template_dict_original[barcode_str_upper]
        # 3. Lowercase string matching
        elif barcode_str and barcode_str.lower() in template_dict_original:
            matched_data = template_dict_original[barcode_str.lower()]
        # 4. Original string matching (exact case)
        elif barcode_str and barcode_str in template_dict_original:
            matched_data = template_dict_original[barcode_str]
        # 5. Try with raw value as string
        elif barcode_raw is not None:
            barcode_raw_str = str(barcode_raw)
            if barcode_raw_str in template_dict_original:
                matched_data = template_dict_original[barcode_raw_str]
            elif barcode_raw_str.upper() in template_dict_original:
                matched_data = template_dict_original[barcode_raw_str.upper()]
            elif barcode_raw_str.lower() in template_dict_original:
                matched_data = template_dict_original[barcode_raw_str.lower()]
        
        # Debug first few rows
        if row_idx <= 6:
            print(f"\nDEBUG Row {row_idx}:")
            print(f"  Raw: '{barcode_raw}' (type: {type(barcode_raw)})")
            print(f"  Str: '{barcode_str}'")
            print(f"  Normalized: '{barcode_normalized}'")
            print(f"  Upper: '{barcode_str_upper}'")
            print(f"  In template_dict: {barcode_normalized in template_dict if barcode_normalized else False}")
            print(f"  In template_dict_original (upper): {barcode_str_upper in template_dict_original if barcode_str_upper else False}")
            print(f"  In template_dict_original (exact): {barcode_str in template_dict_original if barcode_str else False}")
            print(f"  Matched: {matched_data is not None}")
        
        # Debug specific value
        if barcode_str and "TFCLI42021" in barcode_str.upper():
            print(f"\nDEBUG TFCLI42021 - Row {row_idx}:")
            print(f"  Raw: '{barcode_raw}' (type: {type(barcode_raw)})")
            print(f"  Str: '{barcode_str}'")
            print(f"  Normalized: '{barcode_normalized}'")
            print(f"  Upper: '{barcode_str_upper}'")
            print(f"  In template_dict: {barcode_normalized in template_dict if barcode_normalized else False}")
            print(f"  In template_dict_original (upper): {barcode_str_upper in template_dict_original if barcode_str_upper else False}")
            print(f"  In template_dict_original (exact): {barcode_str in template_dict_original if barcode_str else False}")
            print(f"  Matched: {matched_data is not None}")
        
        if matched_data:
            matched_count += 1
            matched_items.append(barcode_normalized or barcode_str)
            # Fill in the data for each new column using the correct column index
            for col_name in columns_to_append:
                if col_name in column_name_to_index:
                    col_idx = column_name_to_index[col_name]
                    value = matched_data.get(col_name, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
        elif barcode_normalized or (barcode_str and barcode_str.lower() not in ['nan', 'none', '']):
            unmatched_items.append((row_idx, barcode_raw, barcode_normalized, barcode_str))
    
    print(f"\nMatched and filled {matched_count} rows out of {ws.max_row - 1} total rows")
    
    # Show unmatched items for debugging
    if unmatched_items:
        print(f"\nUnmatched items ({len(unmatched_items)}):")
        print("First 10 unmatched items:")
        for item_info in unmatched_items[:10]:
            if len(item_info) == 4:
                row_idx, raw, normalized, original_str = item_info
                print(f"  Row {row_idx}: Raw='{raw}' -> Normalized='{normalized}' -> OriginalStr='{original_str}'")
            else:
                row_idx, raw, normalized = item_info
                print(f"  Row {row_idx}: Raw='{raw}' -> Normalized='{normalized}'")
        
        # Show which template values are not being matched
        template_keys_set = set(template_dict.keys())
        template_original_set = set(template_dict_original.keys())
        original_normalized = set()
        original_raw_values = []
        
        for row_idx in range(2, ws.max_row + 1):
            barcode_cell = ws.cell(row=row_idx, column=barcode_col_idx)
            barcode_raw = barcode_cell.value
            barcode_normalized = normalize_item_no(barcode_raw)
            barcode_str = str(barcode_raw).strip() if barcode_raw is not None else ""
            
            if barcode_normalized:
                original_normalized.add(barcode_normalized)
            if barcode_str and barcode_str.lower() not in ['nan', 'none', '']:
                original_raw_values.append(barcode_str)
        
        print(f"\n=== MATCHING ANALYSIS ===")
        print(f"Template normalized keys: {len(template_keys_set)}")
        print(f"Template original keys: {len(template_original_set)}")
        print(f"Original normalized values: {len(original_normalized)}")
        print(f"Original raw values: {len(original_raw_values)}")
        
        # Show intersection
        intersection_normalized = template_keys_set & original_normalized
        print(f"\nIntersection (normalized): {len(intersection_normalized)}")
        if intersection_normalized:
            print(f"  Sample matches: {list(intersection_normalized)[:5]}")
        
        # Check original string matching
        original_raw_set = set(original_raw_values)
        original_raw_upper_set = set(v.upper() for v in original_raw_values)
        intersection_original = template_original_set & original_raw_set
        intersection_original_upper = set(k.upper() for k in template_original_set) & original_raw_upper_set
        
        print(f"Intersection (original exact): {len(intersection_original)}")
        print(f"Intersection (original upper): {len(intersection_original_upper)}")
        
        missing_in_original = template_keys_set - original_normalized
        missing_in_template = original_normalized - template_keys_set
        
        print(f"\n=== SAMPLE VALUES COMPARISON ===")
        print(f"Sample template normalized keys (first 10):")
        for key in list(template_keys_set)[:10]:
            print(f"  '{key}'")
        
        print(f"\nSample original normalized values (first 10):")
        for val in list(original_normalized)[:10]:
            print(f"  '{val}'")
        
        print(f"\nSample template original keys (first 10):")
        for key in list(template_original_set)[:10]:
            print(f"  '{key}'")
        
        print(f"\nSample original raw values (first 10):")
        for val in original_raw_values[:10]:
            print(f"  '{val}'")
        
        if missing_in_template:
            print(f"\nItems in original but not in template ({len(missing_in_template)}):")
            print(f"  Sample: {list(missing_in_template)[:10]}")
        
        if missing_in_original:
            print(f"\nItems in template but not in original ({len(missing_in_original)}):")
            print(f"  Sample: {list(missing_in_original)[:10]}")
    
    # Save the workbook (preserves images)
    print(f"\nSaving merged file to: {OUTPUT_FILE}")
    try:
        # Ensure output directory exists
        OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
        
        # Remove any existing output file to avoid conflicts
        if OUTPUT_FILE.exists():
            try:
                OUTPUT_FILE.unlink()
                print(f"  Removed existing output file")
            except:
                pass
        
        # Save the workbook (openpyxl will preserve what it can)
        print(f"  Saving workbook...")
        wb.save(OUTPUT_FILE)
        print("✓ Done! Data has been merged and saved.")
        print(f"✓ File saved to: {OUTPUT_FILE}")
        
        # Verify file was created and can be opened
        if OUTPUT_FILE.exists():
            file_size = OUTPUT_FILE.stat().st_size / (1024 * 1024)  # Size in MB
            print(f"✓ Output file size: {file_size:.2f} MB")
            
            # Try to verify the file is valid by loading it
            try:
                test_wb = load_workbook(OUTPUT_FILE)
                print(f"✓ File verified - can be opened successfully")
                test_wb.close()
            except Exception as verify_error:
                print(f"⚠️  Warning: File saved but verification failed: {verify_error}")
        else:
            print("❌ Error: Output file was not created!")
            
    except Exception as e:
        print(f"❌ Error saving file: {e}")
        import traceback
        traceback.print_exc()
        print(f"\nTrying to save with minimal options...")
        try:
            # Try saving without any special options
            wb.save(OUTPUT_FILE)
            print(f"✓ Saved successfully with basic options")
        except Exception as e2:
            print(f"❌ Failed to save even with basic options: {e2}")
            return

if __name__ == "__main__":
    main()
